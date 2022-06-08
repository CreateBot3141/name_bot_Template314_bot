
def send_message_table (name,list,user_id,namebot,message_id):
    import iz_telegram
    message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
    markup = ''  

    if 1==1:
        from prettytable import PrettyTable
        mytable = PrettyTable()    
        nm = 0
        for line in list:
            if nm == 0:
                mytable.field_names = line                 
            else:  
                mytable.add_row(line)        
            nm = nm + 1      
        message_out = ''
        message_out = message_out +"<b>"+name+"</b>" + '\n\n'
        message_out = message_out+'<pre>\n'+ str(mytable)+'</pre>\n'        
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)   
        import xlwt
        from datetime import datetime        
        style0 = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',num_format_str='#,##0.00')
        style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        nm = 0
        for line in list:
            if nm == 0:
                print ('[+] line1:',line)
                A1,A2,A3,A4 = line                
                ws['A1'] = A1
                ws['A2'] = A2
                ws['A3'] = A3
                ws['A4'] = A4
                print ('[+] A1:',A1,A2,A3,A4) 
            else:  
                A1,A2,A3,A4 = line
                ws['B1'] = A1
                ws['B2'] = A2
                ws['B3'] = A3
                ws['B4'] = A4
            nm = nm + 1  
        wb.save('sample_bot.xlsx')                
        import telebot
        token = iz_telegram.get_token (namebot)
        bot   = telebot.TeleBot(token)
        doc = open('/home/izofen/Studiya/FL/sample_bot.xlsx', 'rb')
        bot.send_document(user_id,doc)





def start_prog (user_id,namebot,message_in,status,message_id,name_file_picture,telefon_nome,refer,user_id_refer,FIO_id):
    import iz_func
    import iz_game
    import iz_main
    import time
    import iz_telegram
    
    if message_in.find ("_1_") != -1 : 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        name = 'Dashboard accounts'        
        list = iz_func.get_data_to_list ("main","A","D","1","4")
        send_message_table (name,list,user_id,namebot,message_id)



        
    if message_in == 'Пакет 01' or message_in == 'Пакет 02' or message_in == 'Пакет 03' or message_in == 'Пакет 04' or message_in == 'Пакет 05': 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
        markup = ''
        from telebot import types 
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)    

    if message_in == 'FAQ_1_1':
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
        markup = ''  
        name = 'Dashboard accounts'
        field_names  = ["Name", "% portfolio", "Population", "Balance", "Limit %"]
        row1         = ["BTC","0","0,00","0,00000000","50%"]
        row2         = ["ETH","0","0,00","0,00000000","35%"]
        row3         = ["USD","100","1 000 000,00","1 000 000,00","15%"]
        row4         = ["Total","100","1 000 000,00","","100%"]
        from prettytable import PrettyTable
        mytable = PrettyTable()
        mytable.field_names = field_names
        mytable.add_row(row1)
        mytable.add_row(row2)
        mytable.add_row(row3)
        mytable.add_row(row4)        
        message_out = ''
        message_out = message_out +"<b>"+name+"</b>" + '\n\n'
        message_out = message_out+'<pre>\n'+ str(mytable)+'</pre>\n'        
        from telebot import types 
        #answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)    
  
    if message_in == 'New clients_1_1':
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
        markup = ''  
        name = 'Dashboard accounts'
        field_names  = ["Name", "% portfolio", "Population", "Balance", "Limit %"]
        row1         = ["BTC","0","0,00","0,00000000","50%"]
        row2         = ["ETH","0","0,00","0,00000000","35%"]
        row3         = ["USD","100","1 000 000,00","1 000 000,00","15%"]
        row4         = ["Total","100","1 000 000,00","","100%"]
        from prettytable import PrettyTable
        mytable = PrettyTable()
        mytable.field_names = field_names
        mytable.add_row(row1)
        mytable.add_row(row2)
        mytable.add_row(row3)
        mytable.add_row(row4)        
        message_out = ''
        message_out = message_out +"<b>"+name+"</b>" + '\n\n'
        message_out = message_out+'<pre>\n'+ str(mytable)+'</pre>\n'        
        from telebot import types 
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)    
    
    if message_in == 'Big Data clients': 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
        markup = ''  
        name = 'Dashboard accounts'
        field_names  = ["Name", "% portfolio", "Population", "Balance", "Limit %"]
        row1         = ["BTC","0","0,00","0,00000000","50%"]
        row2         = ["ETH","0","0,00","0,00000000","35%"]
        row3         = ["USD","100","1 000 000,00","1 000 000,00","15%"]
        row4         = ["Total","100","1 000 000,00","","100%"]
        from prettytable import PrettyTable
        mytable = PrettyTable()
        mytable.field_names = field_names
        mytable.add_row(row1)
        mytable.add_row(row2)
        mytable.add_row(row3)
        mytable.add_row(row4)        
        message_out = ''
        message_out = message_out +"<b>"+name+"</b>" + '\n\n'
        message_out = message_out+'<pre>\n'+ str(mytable)+'</pre>\n'        
        from telebot import types 
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)    
    
    if message_in == 'Big Data Twitter': 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
        markup = ''  
        name = 'Dashboard accounts'
        field_names  = ["Name", "% portfolio", "Population", "Balance", "Limit %"]
        row1         = ["BTC","0","0,00","0,00000000","50%"]
        row2         = ["ETH","0","0,00","0,00000000","35%"]
        row3         = ["USD","100","1 000 000,00","1 000 000,00","15%"]
        row4         = ["Total","100","1 000 000,00","","100%"]
        from prettytable import PrettyTable
        mytable = PrettyTable()
        mytable.field_names = field_names
        mytable.add_row(row1)
        mytable.add_row(row2)
        mytable.add_row(row3)
        mytable.add_row(row4)        
        message_out = ''
        message_out = message_out +"<b>"+name+"</b>" + '\n\n'
        message_out = message_out+'<pre>\n'+ str(mytable)+'</pre>\n'        
        from telebot import types 
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)    
    
    if message_in == 'Cases': 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
        markup = ''  
        name = 'Dashboard accounts'
        field_names  = ["Name", "% portfolio", "Population", "Balance", "Limit %"]
        row1         = ["BTC","0","0,00","0,00000000","50%"]
        row2         = ["ETH","0","0,00","0,00000000","35%"]
        row3         = ["USD","100","1 000 000,00","1 000 000,00","15%"]
        row4         = ["Total","100","1 000 000,00","","100%"]
        from prettytable import PrettyTable
        mytable = PrettyTable()
        mytable.field_names = field_names
        mytable.add_row(row1)
        mytable.add_row(row2)
        mytable.add_row(row3)
        mytable.add_row(row4)        
        message_out = ''
        message_out = message_out +"<b>"+name+"</b>" + '\n\n'
        message_out = message_out+'<pre>\n'+ str(mytable)+'</pre>\n'        
        from telebot import types 
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)    

    if message_in == 'News': 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
        markup = ''  
        name = 'Dashboard accounts'
        field_names  = ["Name", "% portfolio", "Population", "Balance", "Limit %"]
        row1         = ["BTC","0","0,00","0,00000000","50%"]
        row2         = ["ETH","0","0,00","0,00000000","35%"]
        row3         = ["USD","100","1 000 000,00","1 000 000,00","15%"]
        row4         = ["Total","100","1 000 000,00","","100%"]
        from prettytable import PrettyTable
        mytable = PrettyTable()
        mytable.field_names = field_names
        mytable.add_row(row1)
        mytable.add_row(row2)
        mytable.add_row(row3)
        mytable.add_row(row4)        
        message_out = ''
        message_out = message_out +"<b>"+name+"</b>" + '\n\n'
        message_out = message_out+'<pre>\n'+ str(mytable)+'</pre>\n'        
        from telebot import types 
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)    

    if message_in == 'Events': 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
        markup = ''  
        name = 'Dashboard accounts'
        field_names  = ["Name", "% portfolio", "Population", "Balance", "Limit %"]
        row1         = ["BTC","0","0,00","0,00000000","50%"]
        row2         = ["ETH","0","0,00","0,00000000","35%"]
        row3         = ["USD","100","1 000 000,00","1 000 000,00","15%"]
        row4         = ["Total","100","1 000 000,00","","100%"]
        from prettytable import PrettyTable
        mytable = PrettyTable()
        mytable.field_names = field_names
        mytable.add_row(row1)
        mytable.add_row(row2)
        mytable.add_row(row3)
        mytable.add_row(row4)        
        message_out = ''
        message_out = message_out +"<b>"+name+"</b>" + '\n\n'
        message_out = message_out+'<pre>\n'+ str(mytable)+'</pre>\n'        
        from telebot import types 
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)    

    if message_in == 'Coin and token data': 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
        markup = ''  
        name = 'Dashboard accounts'
        field_names  = ["Name", "% portfolio", "Population", "Balance", "Limit %"]
        row1         = ["BTC","0","0,00","0,00000000","50%"]
        row2         = ["ETH","0","0,00","0,00000000","35%"]
        row3         = ["USD","100","1 000 000,00","1 000 000,00","15%"]
        row4         = ["Total","100","1 000 000,00","","100%"]
        from prettytable import PrettyTable
        mytable = PrettyTable()
        mytable.field_names = field_names
        mytable.add_row(row1)
        mytable.add_row(row2)
        mytable.add_row(row3)
        mytable.add_row(row4)        
        message_out = ''
        message_out = message_out +"<b>"+name+"</b>" + '\n\n'
        message_out = message_out+'<pre>\n'+ str(mytable)+'</pre>\n'        
        from telebot import types 
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)    

    if message_in == 'Market': 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
        markup = ''  
        name = 'Dashboard accounts'
        field_names  = ["Name", "% portfolio", "Population", "Balance", "Limit %"]
        row1         = ["BTC","0","0,00","0,00000000","50%"]
        row2         = ["ETH","0","0,00","0,00000000","35%"]
        row3         = ["USD","100","1 000 000,00","1 000 000,00","15%"]
        row4         = ["Total","100","1 000 000,00","","100%"]
        from prettytable import PrettyTable
        mytable = PrettyTable()
        mytable.field_names = field_names
        mytable.add_row(row1)
        mytable.add_row(row2)
        mytable.add_row(row3)
        mytable.add_row(row4)        
        message_out = ''
        message_out = message_out +"<b>"+name+"</b>" + '\n\n'
        message_out = message_out+'<pre>\n'+ str(mytable)+'</pre>\n'        
        from telebot import types 
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)    

    if message_in == 'Digest':  
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
        markup = ''  
        name = 'Dashboard accounts'
        field_names  = ["Name", "% portfolio", "Population", "Balance", "Limit %"]
        row1         = ["BTC","0","0,00","0,00000000","50%"]
        row2         = ["ETH","0","0,00","0,00000000","35%"]
        row3         = ["USD","100","1 000 000,00","1 000 000,00","15%"]
        row4         = ["Total","100","1 000 000,00","","100%"]
        from prettytable import PrettyTable
        mytable = PrettyTable()
        mytable.field_names = field_names
        mytable.add_row(row1)
        mytable.add_row(row2)
        mytable.add_row(row3)
        mytable.add_row(row4)        
        message_out = ''
        message_out = message_out +"<b>"+name+"</b>" + '\n\n'
        message_out = message_out+'<pre>\n'+ str(mytable)+'</pre>\n'        
        from telebot import types 
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)    
     
    if message_in == 'New clients_1_1':
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Приобрести пакет',namebot)
        markup = ''  
        name = 'Dashboard accounts'
        field_names  = ["Name", "% portfolio", "Population", "Balance", "Limit %"]
        row1         = ["BTC","0","0,00","0,00000000","50%"]
        row2         = ["ETH","0","0,00","0,00000000","35%"]
        row3         = ["USD","100","1 000 000,00","1 000 000,00","15%"]
        row4         = ["Total","100","1 000 000,00","","100%"]
        from prettytable import PrettyTable
        mytable = PrettyTable()
        mytable.field_names = field_names
        mytable.add_row(row1)
        mytable.add_row(row2)
        mytable.add_row(row3)
        mytable.add_row(row4)        
        message_out = ''
        message_out = message_out +"<b>"+name+"</b>" + '\n\n'
        message_out = message_out+'<pre>\n'+ str(mytable)+'</pre>\n'        
        from telebot import types 
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)  

    if message_in == 'Analytics01': 
        name_01 = 'FAQ' 
        name_02 = 'New clients' 
        name_03 = 'Big Data clients' 
        name_04 = 'Big Data Twitter' 
        name_05 = 'Cases' 
        name_06 = 'News' 
        name_07 = 'Events' 
        name_08 = 'Coin and token data' 
        name_09 = 'Market' 
        name_10 = 'Digest'  
        message_out,menu = iz_telegram.get_message (user_id,'Продукты и сервисы вывод',namebot)
        from telebot import types 
        markup = types.InlineKeyboardMarkup(row_width=1)
        mn011 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_01),callback_data = name_01+"_1_1")
        mn012 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_02),callback_data = name_02+"_1_2")
        mn013 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_03),callback_data = name_03+"_1_3")
        mn014 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_04),callback_data = name_04+"_1_4")
        mn015 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_05),callback_data = name_05+"_1_5")
        mn016 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_06),callback_data = name_06+"_1_6 ")  
        mn017 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_07),callback_data = name_07+"_1_7")
        mn018 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_08),callback_data = name_08+"_1_8")
        mn019 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_09),callback_data = name_09+"_1_9")
        mn020 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_10),callback_data = name_10+"_1_10")     
        markup.add(mn011,mn012,mn013,mn014,mn015,mn016,mn017,mn018,mn019,mn020)
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)  
    
    if message_in == 'Trading01': 
        name_01 = 'FAQ' 
        name_02 = 'Bots' 
        message_out,menu = iz_telegram.get_message (user_id,'Продукты и сервисы вывод',namebot)
        from telebot import types 
        markup = types.InlineKeyboardMarkup(row_width=1)
        mn011 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_01),callback_data = name_01+"_2_1")
        mn012 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_02),callback_data = name_02+"_2_2")
        markup.add(mn011,mn012)
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)  
    
    if message_in == 'Investment01': 
        name_01 = 'FAQ' 
        name_02 = 'Bots' 
        name_03 = 'Leads' 
        message_out,menu = iz_telegram.get_message (user_id,'Продукты и сервисы вывод',namebot)
        from telebot import types 
        markup = types.InlineKeyboardMarkup(row_width=1)
        mn011 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_01),callback_data = name_01+"_3_1")
        mn012 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_02),callback_data = name_02+"_3_2")
        mn013 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_02),callback_data = name_03+"_3_3")
        markup.add(mn011,mn012,mn013)
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)  
    
    if message_in == 'AML & Compliance01':
        name_01 = 'FAQ' 
        name_02 = 'Counterparty checking' 
        name_03 = 'Address checking' 
        name_04 = 'Add to blacklist' 
        message_out,menu = iz_telegram.get_message (user_id,'Продукты и сервисы вывод',namebot)
        from telebot import types 
        markup = types.InlineKeyboardMarkup(row_width=1)
        mn011 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_01),callback_data = name_01+"_4_1")
        mn012 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_02),callback_data = name_02+"_4_2")
        mn013 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_03),callback_data = name_03+"_4_3")
        mn014 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_04),callback_data = name_04+"_4_4")
        markup.add(mn011,mn012,mn013,mn014)
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)  
    
    if message_in == 'Tax & Legal01':
        name_01 = 'FAQ' 
        name_02 = 'Tax' 
        name_03 = 'Legal' 

        message_out,menu = iz_telegram.get_message (user_id,'Продукты и сервисы вывод',namebot)
        from telebot import types 
        markup = types.InlineKeyboardMarkup(row_width=1)
        mn011 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_01),callback_data = name_01+"_5_1")
        mn012 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_02),callback_data = name_02+"_5_2")
        mn013 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_03),callback_data = name_03+"_5_3")
        markup.add(mn011,mn012,mn013)
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)  
    
    if message_in == 'Concierge01': 
        name_01 = 'FAQ' 
        name_02 = 'Request' 
        name_03 = 'Private chat' 
        message_out,menu = iz_telegram.get_message (user_id,'Продукты и сервисы вывод',namebot)
        from telebot import types 
        markup = types.InlineKeyboardMarkup(row_width=1)
        mn011 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_01),callback_data = name_01+"_6_1")
        mn012 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_02),callback_data = name_02+"_6_2")
        mn013 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,name_03),callback_data = name_03+"_6_3")
        markup.add(mn011,mn012,mn013)
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)  

    if message_in == 'Продукты и сервисы': 
        message_out,menu = iz_telegram.get_message (user_id,'Продукты и сервисы вывод',namebot)
        from telebot import types 
        markup = types.InlineKeyboardMarkup(row_width=1)
        mn011 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Analytics"),callback_data   = "Analytics01")
        mn012 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Trading"),callback_data     = "Trading01")
        mn013 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Investment"),callback_data  = "Investment01")
        mn014 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"AML & Compliance"),callback_data = "AML & Compliance01")
        mn015 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Tax & Legal"),callback_data = "Tax & Legal01")
        mn016 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Concierge "),callback_data  = "Concierge01")
        markup.add(mn011,mn012,mn013,mn014,mn015,mn016)
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)  
  
    if message_in == 'Пакеты': 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        message_out,menu = iz_telegram.get_message (user_id,'Список пакетов',namebot)
        markup = ''
        from telebot import types 
        markup = types.InlineKeyboardMarkup(row_width=1)
        mn011 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Пакет 01"),callback_data = "Пакет 01")
        mn012 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Пакет 02"),callback_data = "Пакет 02")
        mn013 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Пакет 03"),callback_data = "Пакет 03")
        mn014 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Пакет 04"),callback_data = "Пакет 04")
        mn015 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Пакет 05"),callback_data = "Пакет 05")
        markup.add(mn011,mn012,mn013,mn014,mn015)
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,message_id)    
    
    if message_in == 'Подписка': 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        #message_out,menu,answer = iz_telegram.send_message (user_id,namebot,"Подписка вывод",'S',0) 
        message_out,menu = iz_telegram.get_message (user_id,'Подписка вывод',namebot)
        markup = ''
        from telebot import types 
        markup = types.InlineKeyboardMarkup(row_width=4)
        mn011 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Пакеты"),callback_data = "Пакеты")
        mn012 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Подтвердить платеж"),callback_data = "Подтвердить платеж")
        markup.add(mn011,mn012)
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,0)  
   
    if message_in == 'Профиль': 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        #message_out,menu,answer = iz_telegram.send_message (user_id,namebot,"Подписка вывод",'S',0) 
        message_out,menu = iz_telegram.get_message (user_id,'Профиль вывод',namebot)
        message_out = message_out.replace('%%refer%%',str(refer)) 


        #markup = ''
        #from telebot import types 
        markup = ''
        #markup = types.InlineKeyboardMarkup(row_width=4)
        #mn011 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Пакеты"),callback_data = "Пакеты")
        #mn012 = types.InlineKeyboardButton(text = iz_telegram.get_namekey (user_id,namebot,"Подтвердить платеж"),callback_data = "Подтвердить платеж")
        #markup.add(mn011,mn012)
        answer = iz_telegram.bot_send (user_id,namebot,message_out,markup,0)  
   

    if message_in == '/start': 
        status = '' 
        iz_telegram.save_variable (user_id,namebot,'status','')
        

    if message_in == 'Отмена': 
        status = '' 
        message_out,menu,answer = iz_telegram.send_message (user_id,namebot,"Отмена",'S',0) 
        iz_telegram.save_variable (user_id,namebot,'status','')

  